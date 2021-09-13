import openpyxl
import pandas as pd
import numpy as np
from sklearn.decomposition import PCA
from sklearn.datasets import make_blobs
import xlsxwriter as xlwt
import xlwt

df = pd.read_excel('附件1 近5年402家供应商的相关数据.xlsx', sheet_name='企业的订货量（m³）')
df2 = pd.read_excel('附件1 近5年402家供应商的相关数据.xlsx', sheet_name='供应商的供货量（m³）')

id_list = []
type_list = []

for i in range(1, 10):
    id_list.append('S00' + str(i))

for i in range(10, 100):
    id_list.append('S0' + str(i))

for i in range(100, 403):
    id_list.append('S' + str(i))
print(id_list)

for i in range(402):
    s = df.loc[(df['供应商ID'] == id_list[i])]
    s_list = np.array(s).tolist()[0]
    type_list.append(s_list[1])

# 指标1 订货率
tag_list1 = []

for i in range(402):
    s = df.loc[(df['供应商ID'] == id_list[i])]
    s_list = np.array(s).tolist()[0]
    sum = 0
    for j in range(2, 242):
        if s_list[j] != 0:
            sum += 1
    tag_list1.append(sum / 240)

# 指标2 进步因子
tag_list2 = []

for i in range(402):
    P_list = []
    s = df2.loc[(df['供应商ID'] == id_list[i])]
    s_list = np.array(s).tolist()[0]
    sum = 0
    flag_24 = 0
    for j in range(2, 242):
        sum += s_list[j]
        flag_24 += 1
        if flag_24 == 24:
            P_list.append(sum)
            sum = 0
            flag_24 = 0
    I = 0
    for j in range(0, 9):
        if P_list[j] == 0:
            I += 1
        else:
            I += (P_list[j + 1] - P_list[j]) / P_list[j]
    tag_list2.append(I / 9)

# 指标3 按时交货量率 and 指标6 供货稳定性
tag_list3 = []
tag_list6 = []

for i in range(402):
    s = df.loc[(df['供应商ID'] == id_list[i])]
    s_list = np.array(s).tolist()[0]
    s2 = df2.loc[(df['供应商ID'] == id_list[i])]
    s_list2 = np.array(s2).tolist()[0]
    rate_list = []
    for j in range(2, 242):
        if s_list[j] != 0:
            rate_list.append(s_list2[j] / s_list[j])
    sum = 0
    for j in range(len(rate_list)):
        sum += rate_list[j]
    tag_list3.append(sum / len(rate_list))

    # 计算指标6
    tag_list6.append(np.var(rate_list))

# 指标4 按时交货量率
tag_list4 = []

for i in range(402):
    s = df.loc[(df['供应商ID'] == id_list[i])]
    s_list = np.array(s).tolist()[0]
    s2 = df2.loc[(df['供应商ID'] == id_list[i])]
    s_list2 = np.array(s2).tolist()[0]
    M = 0
    N = 0
    for j in range(2, 242):
        if s_list[j] != 0:
            N += 1
        if s_list[j] != 0 and s_list2[j] == 0:
            M += 1
    tag_list4.append(1 - (M / N))

# 指标5 供货规模
tag_list5 = []

for i in range(402):
    s = df.loc[(df['供应商ID'] == id_list[i])]
    s_list = np.array(s).tolist()[0]
    s2 = df2.loc[(df['供应商ID'] == id_list[i])]
    s_list2 = np.array(s2).tolist()[0]
    sum = 0
    num = 0
    for j in range(2, 242):
        if s_list2[j] != 0:
            num += 1
    quan = 0
    for j in range(num):
        quan +=  j + 1
    flag_0 = 0
    for j in range(2, 242):
        if s_list2[j] == 0:
            flag_0 += 1
        else:
            sum += s_list2[j] * (j - flag_0 - 1) / quan
    tag_list5.append(sum)

# 指标7 单项特长
tag_list7 = []
A_list = []
B_list = []
C_list = []

s = df2.loc[(df['材料分类'] == 'A')]
s_list = np.array(s).tolist()
sum = 0
flag_24 = 0
for i in range(2, 242):
    for j in range(len(s_list)):
        sum += s_list[j][i]
    flag_24 += 1
    if flag_24 == 24:
        A_list.append(sum)
        sum = 0
        flag_24 = 0

s = df2.loc[(df['材料分类'] == 'B')]
s_list = np.array(s).tolist()
sum = 0
flag_24 = 0
for i in range(2, 242):
    for j in range(len(s_list)):
        sum += s_list[j][i]
    flag_24 += 1
    if flag_24 == 24:
        B_list.append(sum)
        sum = 0
        flag_24 = 0

s = df2.loc[(df['材料分类'] == 'C')]
s_list = np.array(s).tolist()
sum = 0
flag_24 = 0
for i in range(2, 242):
    for j in range(len(s_list)):
        sum += s_list[j][i]
    flag_24 += 1
    if flag_24 == 24:
        C_list.append(sum)
        sum = 0
        flag_24 = 0

for i in range(402):
    tens = []
    s = df2.loc[(df['供应商ID'] == id_list[i])]
    s_list = np.array(s).tolist()[0]
    sum = 0
    flag_24 = 0
    for j in range(2, 242):
        sum += s_list[j]
        flag_24 += 1
        if flag_24 == 24:
            if s_list[1] == 'A':
                tens.append(sum / A_list[int((j - 2) / 24)])
            elif s_list[1] == 'B':
                tens.append(sum / B_list[int((j - 2) / 24)])
            else:
                tens.append(sum / C_list[int((j - 2) / 24)])
            sum = 0
            flag_24 = 0
    tag_list7.append(np.mean(tens))

# 主成分分析
std_tag_list = []
temp = []
ave_list = []
r_list = []
tag_list = [tag_list1, tag_list2, tag_list3, tag_list4, tag_list5, tag_list6, tag_list7]
for i in range(7):
    ave = np.mean(tag_list[i])
    ave_list.append(ave)
    r = np.std(tag_list[i], ddof=1)
    r_list.append(r)
    for j in range(402):
        temp.append((tag_list[i][j] - ave) / r)
    std_tag_list.append(temp)
    temp = []

b = []
for i in range(402):
    temp = []
    for j in range(7):
        temp.append(std_tag_list[j][i])
    b.append(temp)
c = np.array(b)

md = PCA().fit(c)
# print("特征值", md.explained_variance_)
# print("贡献率", md.explained_variance_ratio_)
# print("系数", md.components_)

k_list = [0, 0, 0, 0, 0, 0, 0]
for i in range(4):
    xishu = (md.explained_variance_ratio_[i] / (md.explained_variance_ratio_[0] + md.explained_variance_ratio_[1] +
                                                md.explained_variance_ratio_[2] + md.explained_variance_ratio_[3]))
    for j in range(7):
        k_list[j] += xishu * md.components_[i][j]

# 接下来计算402家供货商得分情况
score_list = []
for i in range(402):
    single_score = 0
    for j in range(7):
        single_score += k_list[j] * std_tag_list[j][i]
    score_list.append({"shop": id_list[i], "type": type_list[i], "score": single_score})

for i in range(402):
    for j in range(401):
        if score_list[j].get("score") < score_list[j + 1].get("score"):
            temp = score_list[j]
            score_list[j] = score_list[j + 1]
            score_list[j + 1] = temp

final_list = []

for i in range(402):
    final_list.append({"shop": score_list[i].get("shop"), "type": score_list[i].get("type"),
                       "score": (score_list[i].get("score") -
                                 score_list[401].get("score")) /
                                (score_list[0].get("score") -
                                 score_list[401].get("score"))})

xl = openpyxl.Workbook()
xl.active
sheet = xl.create_sheet("score")
sheet.cell(1, 1, '排名')
sheet.cell(1, 2, '供应商ID')
sheet.cell(1, 3, '材料分类')
sheet.cell(1, 4, '得分情况')
for i in range(50):
    sheet.cell(i + 1 + 1, 1, i + 1)
    sheet.cell(i + 1 + 1, 2, final_list[i].get("shop"))
    sheet.cell(i + 1 + 1, 3, final_list[i].get("type"))
    sheet.cell(i + 1 + 1, 4, final_list[i].get("score"))
xl.save('rank.xlsx')

